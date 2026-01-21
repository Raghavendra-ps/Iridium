from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
import pandas as pd
import io
import json
from typing import List, Dict, Any, Optional
import uuid
from pathlib import Path

from app.api import dependencies
from app.core.services import job_service
from app.db.models import User, LinkedOrganization
from app.db.session import get_db
from app.infrastructure.erpnext_client import ERPNextClient

router = APIRouter()

UPLOAD_DIR = Path("/app/uploads")
PROCESSED_DIR = UPLOAD_DIR / "processed"

class SheetInfoRequest(BaseModel):
    target_org_id: int

class SheetDataSubmission(BaseModel):
    year: int
    month: int
    data: List[Dict[str, Any]]
    target_org_id: Optional[int] = None

@router.post("/get-employees-for-sheet")
async def get_employees_for_sheet(
    *,
    db: Session = Depends(get_db),
    # --- START OF FIX: Use the correct dependency name ---
    current_user: User = Depends(dependencies.get_current_internal_user),
    # --- END OF FIX ---
    request_data: SheetInfoRequest,
):
    """Fetches all active employees from a target organization to populate the sheet maker grid."""
    org = db.query(LinkedOrganization).filter(
        LinkedOrganization.id == request_data.target_org_id, 
    ).first()
    # Note: Security is implicitly handled because only internal users can access this.
    # A future improvement could be to check if the org is assigned to a manager's company.
    if not org:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Target organization not found.")

    try:
        erp_client = ERPNextClient(base_url=org.erpnext_url, api_key=org.api_key, api_secret=org.api_secret)
        erp_employees = await erp_client.get_all_employees()
        return sorted(
            [{"employee_code": emp.get("company_employee_id") or emp.get("name"), 
              "employee_name": emp.get("employee_name")} for emp in erp_employees],
            key=lambda x: x['employee_name']
        ) if erp_employees else []
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail=f"Could not connect to ERPNext: {e}")


@router.post("/create-sheet-from-data")
def create_sheet_from_data(
    *,
    # --- START OF FIX: Use the correct dependency name ---
    current_user: User = Depends(dependencies.get_current_internal_user),
    # --- END OF FIX ---
    submission_data: SheetDataSubmission,
):
    """Receives grid data, converts it to Excel, and returns for download."""
    if not submission_data.data:
        raise HTTPException(status_code=400, detail="No data provided to create sheet.")
    
    df = pd.DataFrame(submission_data.data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Attendance')
        worksheet = writer.sheets['Attendance']
        
        from openpyxl.utils import get_column_letter
        for idx, col_name in enumerate(df.columns):
            max_len = max(df[col_name].astype(str).map(len).max(), len(str(col_name))) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len
    
    output.seek(0)
    filename = f"Generated_Attendance_{submission_data.year}-{submission_data.month:02d}.xlsx"
    return StreamingResponse(
        output, 
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@router.post("/save-as-job", status_code=status.HTTP_201_CREATED)
def save_sheet_as_job(
    *,
    db: Session = Depends(get_db),
    # This can be used by any active user (manager, employee, admin)
    current_user: User = Depends(dependencies.get_current_active_user),
    submission_data: SheetDataSubmission,
):
    """
    Receives grid data from the sheet maker, saves it as a JSON file,
    and creates a new ConversionJob for the user.
    """
    if not submission_data.data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No data provided to save.")
    
    # In the new model, jobs are linked to the user's main organization.
    # The target_org_id refers to the ERPNext-linked org for submission.
    if not submission_data.target_org_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="A target organization is required to save a job.")

    year = submission_data.year
    month = submission_data.month
    
    records = []
    for row in submission_data.data:
        emp_code = row.get("Employee Code")
        emp_name = row.get("Employee Name")
        if not emp_code:
            continue
        for day, status_value in row.items():
            if day.isdigit() and status_value:
                date = f"{year}-{month:02d}-{int(day):02d}"
                records.append({
                    "employee": emp_code, 
                    "employee_name": emp_name,
                    "attendance_date": date, 
                    "status": str(status_value)
                })

    if not records:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No attendance data to save.")

    PROCESSED_DIR.mkdir(exist_ok=True, parents=True)
    job_uuid = uuid.uuid4()
    processed_filename = f"{job_uuid}_processed.json"
    processed_path = PROCESSED_DIR / processed_filename
    
    with open(processed_path, "w") as f:
        json.dump(records, f, indent=2)

    filename = f"Generated Sheet - {year}-{month:02d}.xlsx"
    
    job = job_service.create_job(
        db=db, 
        owner_id=current_user.id, 
        original_filename=filename,
        storage_filename=f"{job_uuid}.json", 
        target_doctype="attendance",
        target_org_id=submission_data.target_org_id,
        # We need to add attendance_year and attendance_month here
        attendance_year=year,
        attendance_month=month
    )
    
    job.processed_data_path = str(processed_path)
    # This job is pre-validated by the user, so it can go straight to AWAITING_VALIDATION
    job.status = "AWAITING_VALIDATION" 
    db.commit()
    db.refresh(job)

    return job